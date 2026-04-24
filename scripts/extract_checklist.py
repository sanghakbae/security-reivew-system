from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path("/Users/mac/Downloads/BS-IS-PRO-003_보안성+검토+체크리스트.xlsx")
ROOT = Path(__file__).resolve().parents[1]


def q(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def main() -> None:
    workbook = load_workbook(SOURCE, data_only=True)
    sheet = workbook["2. 보안성 검토 요구사항"]
    rows = []

    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row[1]:
            continue
        rows.append(
            {
                "code": str(row[1]).strip(),
                "category": str(row[2]).strip().replace("\n", " ") if row[2] else "",
                "title": str(row[3]).strip() if row[3] else "",
                "requirement": str(row[4]).strip() if row[4] else "",
                "applies_personal": bool(row[5]),
                "applies_non_personal": bool(row[6]),
                "description": str(row[7]).strip() if row[7] else "",
                "sort_order": len(rows) + 1,
            }
        )

    data_file = ROOT / "src" / "data" / "checklist.ts"
    data_file.parent.mkdir(parents=True, exist_ok=True)
    data_file.write_text(
        "import type { SecurityRequirement } from \"../types\";\n\n"
        "export const CHECKLIST: SecurityRequirement[] = "
        + json.dumps(rows, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )

    seed_file = ROOT / "supabase" / "seed_checklist.sql"
    seed_file.parent.mkdir(parents=True, exist_ok=True)
    values = []
    for item in rows:
        values.append(
            "("
            + ", ".join(
                [
                    q(item["code"]),
                    q(item["category"]),
                    q(item["title"]),
                    q(item["requirement"]),
                    q(item["description"]),
                    "true" if item["applies_personal"] else "false",
                    "true" if item["applies_non_personal"] else "false",
                    str(item["sort_order"]),
                ]
            )
            + ")"
        )

    seed_file.write_text(
        "insert into public.sr_security_requirements "
        "(code, category, title, requirement, description, applies_personal, applies_non_personal, sort_order)\n"
        "values\n"
        + ",\n".join(values)
        + "\non conflict (code) do update set\n"
        "  category = excluded.category,\n"
        "  title = excluded.title,\n"
        "  requirement = excluded.requirement,\n"
        "  description = excluded.description,\n"
        "  applies_personal = excluded.applies_personal,\n"
        "  applies_non_personal = excluded.applies_non_personal,\n"
        "  sort_order = excluded.sort_order;\n",
        encoding="utf-8",
    )

    print(f"Extracted {len(rows)} checklist requirements")


if __name__ == "__main__":
    main()
